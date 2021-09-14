import bs4
import urllib.request
import openpyxl
import os
from openpyxl import load_workbook

#playerid bug
#연도 입력
year = str(2019)
club_idt = [0 for i in range(40)]
club_idt[0] = '32122'#한체대
club_idt[1] = '33835'#가천대
club_idt[2] = '23604'#가톨릭대
club_idt[3] = '11841'#건대글로컬
club_idt[4] = '11866'#건대서울
club_idt[5] = '24037'#경기대
club_idt[6] = '23749'#경희국제
club_idt[7] = '23553'#경희서울
club_idt[8] = '11853'#고려대
club_idt[9] = '23999'#광운대
club_idt[10] = '11714'#국민대
club_idt[11] = '21584'#단국죽전
club_idt[12] = '11869'#단국천안
club_idt[13] = '23987'#동국대
club_idt[14] = '11849'#명지서울
club_idt[15] = '23995'#명지용인
club_idt[16] = '25226'#백석대
club_idt[17] = '23515'#상명대
club_idt[18] = '11721'#서강대
club_idt[19] = '23686'#서경대
club_idt[20] = '970'#서울과학기술대
club_idt[21] = '11709'#서울시립대
club_idt[22] = '24158'#성균관대
club_idt[23] = '11810'#세종대
club_idt[24] = '11877'#숭실대
club_idt[25] = '343'#아주대
club_idt[26] = '11919'#연세대
club_idt[27] = '11847'#외대글로벌
club_idt[28] = '23676'#인천대
club_idt[29] = '471'#인하대
club_idt[30] = '24016'#중앙대
club_idt[31] = '28867'#한국교통대
club_idt[32] = '6205'#산기대
club_idt[33] = '23728'#한국외대서울
club_idt[34] = '23593'#항공대
club_idt[35] = '24010'#한성대
club_idt[36] = '23691'#한신대
club_idt[37] = '24012'#한양대에리카
club_idt[38] = '23961'#한양대서울
club_idt[39] = '23511'#홍익대

accumulate_sum=0

#엑셀 파일 관련
#import openpyxl
excel_file = openpyxl.Workbook()
excel_file.remove(excel_file.active)
excel_sheet = excel_file.create_sheet('선수 이름')
excel_sheet = excel_file.create_sheet('팀 기록1') # 조회시 에러 방지
sheet = excel_file.active

for teams in range(0, len(club_idt)):
        club_id = club_idt[teams]
#for teams in range(23, 24):
        #club_id = club_idt[teams]
        url = "http://www.gameone.kr/club/info/player?club_idx="+ club_id# 한체대 주소
        #향후 32122 이 부분을 한체대에 배정할 예정 ex) 한체대 = 32122
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")

        #팀 이름
        team_name = bsObj.select('title') # team_name[0].text 로 사용


        #선수 이름 주소
        player_name = bsObj.select('form > dt')

        #선수 이름 양식
        player_array = [0 for i in range(len(player_name))]

        count=0
        for num in range(0, len(player_name)):
                for back_num in range(0,1):
                        player_array[count] = player_name[num].text.split('.')
                        count+=1
        #이름만 저장
        name = [0 for i in range(len(player_name))]
        for num in range(0, len(player_array)):
                name[num] = player_array[num][1]

        #등번호만 저장
        backnum = [0 for i in range(len(player_name))]
        for num in range(0, len(player_array)):
                backnum[num] = player_array[num][0]

        #기록 형식에 맞게 이름 저장 윤주엽(5)
        #player = [0 for i in range(len(player_name))]
        #for num in range(0, len(player_array)):
        #        player[num] = name[num] + '(' + backnum[num] + ')'


        #여러 엑셀 파일 읽기
        #import os
        #경로 지정
        #path = "/Users/YEO/AppData/Local/Programs/Python/Python38/"+ year + "/" + team_name[0].text + "/"
        path = "/Users/JuYeop/AppData/Local/Programs/Python/Python38-32/"+ year + "/" + team_name[0].text + "/"
        file_list = os.listdir(path)
        
        ###투수 기록을 위해 변수 선언
        Player_Pitcher_Stats = [0 for i in range(len(name))]

        for i in range(0, len(name)):
                Player_Pitcher_Stats[i] = dict.fromkeys({'playerId','G', 'Win', 'Lose', 'HLD', 'SV', 'IP', 'Batter', 'AB', 'H','HR', 'SAC', 'SF',
                                                         'BB', 'HBP', 'K', 'WP', 'BK', 'R', 'ER', 'PI', 'ERA'},0)
        
        if(os.path.exists(path) != True): # 기록 없으면 continue
                print(team_name[0].text + " (" + club_id + ") 기록 없음")
                continue
        

        #폴더 조회 후 없으면 만들기
        if(os.path.exists(path) != True): # 폴더가 존재하지 않는다면
                os.makedirs(path) # 폴더를 만든다.

        #누적 기록 저장소
        accumulate_record = [0 for i in range(len(name))]
        for i in range(0, len(name)):
                accumulate_record[i] = []




        ###기록 가져오기#
        #from openpyxl import load_workbook

        for repeat in range(0, len(file_list)):
                load_wb = load_workbook(path + file_list[repeat], data_only=True)
                load_ws = load_wb['팀 기록1']

                #선수 기록 저장
                player_record = [0 for i in range(len(name))]
                for i in range(0, len(name)):
                        player_record[i] = []
                        
                ################
                #홈 타자 수 + 원정 타자 수 + 1, '이닝' 나올때까지 찾는 원리
                batter_number=0

                for i in range(0, len(name)):
                        if(load_ws.cell(7+i,3).value != '이닝'):
                                batter_number+=1
                        else:
                                break
                
                Pitcher_count = 7+batter_number+1
                Pitcher_away=0
                Pitcher_home=0
                
                #원정 투수 숫자 세기
                for i in range(0, len(name)):
                        if(load_ws.cell(Pitcher_count+i,1).value != None):
                                Pitcher_away+=1
                        else:
                                break
                #홈 투수 숫자 세기
                for i in range(0, len(name)):
                        if(load_ws.cell(Pitcher_count+Pitcher_away+1+i,1).value != None):
                                Pitcher_home+=1
                        else:
                                break         
                #전체 투수 숫자 세기
                Pitcher_all = Pitcher_away + 1 + Pitcher_home

                #홈 원정 구분(동명이인 방지)
                
                # 확인하는 팀이 원정팀
                if(load_ws.cell(Pitcher_count-2,2).value == ' ' +team_name[0].text):
                        
                        #출장 게임수 넣기
                        for i in range(0, Pitcher_away):
                            for j in range(0, len(name)):
                                if(load_ws.cell(Pitcher_count+i, 1).value == name[j]):
                                    Player_Pitcher_Stats[j]['G'] +=1
                                    break

                        #승, 패, 세, 홀 기록
                        for i in range(0, Pitcher_away): #전체선수숫자조회
                            for j in range(0, len(name)):
                                if(load_ws.cell(Pitcher_count+i, 1).value == name[j]):                        
                                            if(load_ws.cell(Pitcher_count+i, 2).value == '승'):
                                                Player_Pitcher_Stats[j]['Win'] +=1
                                            elif(load_ws.cell(Pitcher_count+i, 2).value == '패'):
                                                Player_Pitcher_Stats[j]['Lose'] +=1
                                            elif(load_ws.cell(Pitcher_count+i, 2).value == '홀'):
                                                Player_Pitcher_Stats[j]['HLD'] +=1
                                            elif(load_ws.cell(Pitcher_count+i, 2).value == '세'):
                                                Player_Pitcher_Stats[j]['SV'] +=1
                                            elif(load_ws.cell(Pitcher_count+i, 2).value != '-'):
                                                print(load_ws.cell(Pitcher_count+j, 2).value)

                        #나머지 투수 기록
                        for i in range(0, Pitcher_away): #전체선수숫자조회
                            for j in range(0, len(name)):
                                if(load_ws.cell(Pitcher_count+i, 1).value == name[j]):                        
                                            if(load_ws.cell(Pitcher_count+i, 3).value != None):
                                                if '⅓' in load_ws.cell(Pitcher_count+i, 3).value:
                                                        Outcount = float(load_ws.cell(Pitcher_count+i, 3).value.split(' ')[0])*3 +float(load_ws.cell(Pitcher_count+i, 3).value.split(' ')[1].replace("⅓","1"))
                                                elif '⅔' in load_ws.cell(Pitcher_count+i, 3).value:
                                                        Outcount = float(load_ws.cell(Pitcher_count+i, 3).value.split(' ')[0])*3 +float(load_ws.cell(Pitcher_count+i, 3).value.split(' ')[1].replace("⅔","2"))
                                                else:
                                                        Outcount = float(load_ws.cell(Pitcher_count+i, 3).value.split(' ')[0])*3   
                                                Player_Pitcher_Stats[j]['IP'] += float(Outcount) # 아웃카운트 기준, 이닝= /3
                                                Player_Pitcher_Stats[j]['Batter'] += int(load_ws.cell(Pitcher_count+i, 4).value)
                                                Player_Pitcher_Stats[j]['AB'] += int(load_ws.cell(Pitcher_count+i, 5).value)
                                                Player_Pitcher_Stats[j]['H'] += int(load_ws.cell(Pitcher_count+i, 6).value)
                                                Player_Pitcher_Stats[j]['HR'] += int(load_ws.cell(Pitcher_count+i, 7).value)
                                                Player_Pitcher_Stats[j]['SAC'] += int(load_ws.cell(Pitcher_count+i, 8).value)
                                                Player_Pitcher_Stats[j]['SF'] += int(load_ws.cell(Pitcher_count+i, 9).value)
                                                Player_Pitcher_Stats[j]['BB'] += int(load_ws.cell(Pitcher_count+i, 10).value)
                                                Player_Pitcher_Stats[j]['HBP'] += int(load_ws.cell(Pitcher_count+i, 11).value)
                                                Player_Pitcher_Stats[j]['K'] += int(load_ws.cell(Pitcher_count+i, 12).value)
                                                Player_Pitcher_Stats[j]['WP'] += int(load_ws.cell(Pitcher_count+i, 13).value)
                                                Player_Pitcher_Stats[j]['BK'] += int(load_ws.cell(Pitcher_count+i, 14).value)
                                                Player_Pitcher_Stats[j]['R'] += int(load_ws.cell(Pitcher_count+i, 15).value)
                                                Player_Pitcher_Stats[j]['ER'] += int(load_ws.cell(Pitcher_count+i, 16).value)
                                                Player_Pitcher_Stats[j]['PI'] += int(load_ws.cell(Pitcher_count+i, 17).value)

                # 확인하는 팀이 홈팀
                elif(load_ws.cell(Pitcher_count+Pitcher_away,2).value == ' ' +team_name[0].text):
                           
                        #출장 게임수 넣기
                        for i in range(0, Pitcher_home):
                            for j in range(0, len(name)):
                                if(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 1).value == name[j]):
                                    Player_Pitcher_Stats[j]['G'] +=1
                                    break

                        #승, 패, 세, 홀 기록
                        for i in range(0, Pitcher_home): #전체선수숫자조회
                            for j in range(0, len(name)):
                                if(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 1).value == name[j]):                        
                                            if(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 2).value == '승'):
                                                Player_Pitcher_Stats[j]['Win'] +=1
                                            elif(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 2).value == '패'):
                                                Player_Pitcher_Stats[j]['Lose'] +=1
                                            elif(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 2).value == '홀'):
                                                Player_Pitcher_Stats[j]['HLD'] +=1
                                            elif(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 2).value == '세'):
                                                Player_Pitcher_Stats[j]['SV'] +=1
                                            elif(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 2).value != '-'):
                                                print(load_ws.cell(Pitcher_count+j, 2).value)

                        #나머지 투수 기록
                        for i in range(0, Pitcher_home): #전체선수숫자조회
                            for j in range(0, len(name)):
                                if(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 1).value == name[j]):                        
                                            if(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value != None):
                                                if '⅓' in load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value:
                                                        Outcount = float(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value.split(' ')[0])*3 +float(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value.split(' ')[1].replace("⅓","1"))
                                                elif '⅔' in load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value:
                                                        Outcount = float(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value.split(' ')[0])*3 +float(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value.split(' ')[1].replace("⅔","2"))
                                                else:
                                                        Outcount = float(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 3).value.split(' ')[0])*3   
                                                Player_Pitcher_Stats[j]['IP'] += float(Outcount) # 아웃카운트 기준, 이닝= /3
                                                Player_Pitcher_Stats[j]['Batter'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 4).value)
                                                Player_Pitcher_Stats[j]['AB'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 5).value)
                                                Player_Pitcher_Stats[j]['H'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 6).value)
                                                Player_Pitcher_Stats[j]['HR'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 7).value)
                                                Player_Pitcher_Stats[j]['SAC'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 8).value)
                                                Player_Pitcher_Stats[j]['SF'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 9).value)
                                                Player_Pitcher_Stats[j]['BB'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 10).value)
                                                Player_Pitcher_Stats[j]['HBP'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 11).value)
                                                Player_Pitcher_Stats[j]['K'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 12).value)
                                                Player_Pitcher_Stats[j]['WP'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 13).value)
                                                Player_Pitcher_Stats[j]['BK'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 14).value)
                                                Player_Pitcher_Stats[j]['R'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 15).value)
                                                Player_Pitcher_Stats[j]['ER'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 16).value)
                                                Player_Pitcher_Stats[j]['PI'] += int(load_ws.cell(Pitcher_count+i+Pitcher_away+1, 17).value)


        teamId = 0
        #teamId 입력받기
        if(club_id == '32122'): #한체대
                teamId = '1'
        elif(club_id == '33835'): # 가천대
                teamId = '2'
        elif(club_id == '23604'): # 가톨릭대
                teamId = '3'     
        elif(club_id == '11841'): # 건대(글로컬)
                teamId = '4'
        elif(club_id == '11866'): # 건대(서울)
                teamId = '5'
                
        elif(club_id == '24037'): # 경기대
                teamId = '6'
        elif(club_id == '23749'): # 경희대(국제)
                teamId = '7'
        elif(club_id == '23553'): # 경희대(서울)
                teamId = '8'
        elif(club_id == '11853'): # 고려대
                teamId = '9'
        elif(club_id == '23999'): # 광운대
                teamId = '10'
                
        elif(club_id == '11714'): # 국민대
                teamId = '11'
        elif(club_id == '21584'): # 단국대(죽전)
                teamId = '12'
        elif(club_id == '11869'): # 단국대(천안)
                teamId = '13'
        elif(club_id == '23987'): # 동국대
                teamId = '14'
        elif(club_id == '11849'): # 명지대(서울)
                teamId = '15'
                
        elif(club_id == '23995'): # 명지대(용인)
                teamId = '16'
        elif(club_id == '25226'): # 백석대
                teamId = '17'
        elif(club_id == '23515'): # 상명대
                teamId = '18'
        elif(club_id == '11721'): # 서강대
                teamId = '19'
        elif(club_id == '23686'): # 서경대
                teamId = '20'
                
        elif(club_id == '970'): # 서울과학기술대
                teamId = '21'
        elif(club_id == '11709'): # 서울시립대
                teamId = '22'
        elif(club_id == '24158'): # 성균관대
                teamId = '23'
        elif(club_id == '11810'): # 세종대
                teamId = '24'
        elif(club_id == '11877'): # 숭실대
                teamId = '25'
                
        elif(club_id == '343'): # 아주대
                teamId = '26'
        elif(club_id == '11919'): # 연세대
                teamId = '27'
        elif(club_id == '11847'): # 외대(글로벌)
                teamId = '28'
        elif(club_id == '23676'): # 인천대
                teamId = '29'
        elif(club_id == '471'): # 인하대
                teamId = '30'
                
        elif(club_id == '24016'): # 중앙대
                teamId = '31'
        elif(club_id == '28867'): # 한국교통대
                teamId = '32'
        elif(club_id == '6205'): # 한국산기대
                teamId = '33'
        elif(club_id == '23728'): # 한국외대(서울)
                teamId = '34'
        elif(club_id == '23593'): # 한국항공대
                teamId = '35'
                
        elif(club_id == '24010'): # 한성대
                teamId = '36'
        elif(club_id == '23691'): # 한신대
                teamId = '37'
        elif(club_id == '24012'): # 한양대(에리카)
                teamId = '38'
        elif(club_id == '23961'): # 한양대(서울)
                teamId = '39'
        elif(club_id == '23511'): # 홍익대
                teamId = '40'

        #playerId 입력받기
        #path2 = "/Users/YEO/AppData/Local/Programs/Python/Python38/"
        path2 = "/Users/JuYeop/AppData/Local/Programs/Python/Python38-32/"
        file = load_workbook(path2 + "player_info.xlsx", data_only = True)
        file_open = file['Sheet1']

        #행 갯수 세기
        row_count=0
        for row in file_open.rows:
                row_count+=1

        row_count_str = str(row_count)

        for i in range(1, row_count+1):
                if(str(file_open['B'+str(i)].value) == teamId): #학교 탐색
                        for j in range(0, len(name)):
                                if(name[j] == file_open['D'+str(i)].value): # 이름 탐색
                                        Player_Pitcher_Stats[j]['playerId'] = file_open['A'+str(i)].value
                                        break

        #연도, 이름 양식 입력
        sheet.cell(row = 1, column = 1).value = '선수명'
        sheet.cell(row = 1, column = 2).value = '등번호'
        sheet.cell(row = 1, column = 3).value = 'playerId'
        sheet.cell(row = 1, column = 4).value = 'year'
        sheet.cell(row = 1, column = 5).value = 'G'
        sheet.cell(row = 1, column = 6).value = 'Win'
        sheet.cell(row = 1, column = 7).value = 'Lose'
        sheet.cell(row = 1, column = 8).value = 'HLD'
        sheet.cell(row = 1, column = 9).value = 'SV'
        sheet.cell(row = 1, column = 10).value = 'IP'
        sheet.cell(row = 1, column = 11).value = 'Batter'
        sheet.cell(row = 1, column = 12).value = 'AB'
        sheet.cell(row = 1, column = 13).value = 'H'
        sheet.cell(row = 1, column = 14).value = 'HR'
        sheet.cell(row = 1, column = 15).value = 'SAC'
        sheet.cell(row = 1, column = 16).value = 'SF'
        sheet.cell(row = 1, column = 17).value = 'BB'
        sheet.cell(row = 1, column = 18).value = 'HBP'
        sheet.cell(row = 1, column = 19).value = 'K'
        sheet.cell(row = 1, column = 20).value = 'WP'
        sheet.cell(row = 1, column = 21).value = 'BK'
        sheet.cell(row = 1, column = 22).value = 'R'
        sheet.cell(row = 1, column = 23).value = 'ER'
        sheet.cell(row = 1, column = 24).value = 'PI'
        sheet.cell(row = 1, column = 25).value = 'ERA'
        sheet.cell(row = 1, column = 26).value = 'FIP'
        sheet.cell(row = 1, column = 27).value = 'OPS'
        sheet.cell(row = 1, column = 28).value = 'WAR'
        sheet.cell(row = 1, column = 29).value = 'GO'
        sheet.cell(row = 1, column = 30).value = 'FO'
        sheet.cell(row = 1, column = 31).value = 'LO'
        sheet.cell(row = 1, column = 32).value = 'P'
        sheet.cell(row = 1, column = 33).value = 'C'
        sheet.cell(row = 1, column = 34).value = '1B'
        sheet.cell(row = 1, column = 35).value = '2B'
        sheet.cell(row = 1, column = 36).value = '3B'
        sheet.cell(row = 1, column = 37).value = 'SS'
        sheet.cell(row = 1, column = 38).value = 'LF'
        sheet.cell(row = 1, column = 39).value = 'LC'
        sheet.cell(row = 1, column = 40).value = 'CF'
        sheet.cell(row = 1, column = 41).value = 'RC'
        sheet.cell(row = 1, column = 42).value = 'RF'


        #기록 양식 저장
        num=0
        #Player_Batter_Stats[0].keys()

        #for i in range(0, len(name)):
        #        sheet.cell(row = 1, column = 3+num) = Player_Batter_Stats[i]['playerId']
        #        num+=1
        #기록 엑셀에 저장
        
        
        for i in range(0, len(name)):
                sheet.cell(row = 2+i+accumulate_sum, column = 1).value = name[i] # 선수명 기입
                sheet.cell(row = 2+i+accumulate_sum, column = 2).value = backnum[i] # 등번호 기입
                sheet.cell(row = 2+i+accumulate_sum, column = 3).value = Player_Pitcher_Stats[i]['playerId']
                sheet.cell(row = 2+i+accumulate_sum, column = 4).value = year
                sheet.cell(row = 2+i+accumulate_sum, column = 5).value = Player_Pitcher_Stats[i]['G']
                sheet.cell(row = 2+i+accumulate_sum, column = 6).value = Player_Pitcher_Stats[i]['Win']
                sheet.cell(row = 2+i+accumulate_sum, column = 7).value = Player_Pitcher_Stats[i]['Lose']
                sheet.cell(row = 2+i+accumulate_sum, column = 8).value = Player_Pitcher_Stats[i]['HLD']
                sheet.cell(row = 2+i+accumulate_sum, column = 9).value = Player_Pitcher_Stats[i]['SV']
                sheet.cell(row = 2+i+accumulate_sum, column = 10).value = Player_Pitcher_Stats[i]['IP']
                sheet.cell(row = 2+i+accumulate_sum, column = 11).value = Player_Pitcher_Stats[i]['Batter']
                sheet.cell(row = 2+i+accumulate_sum, column = 12).value = Player_Pitcher_Stats[i]['AB']
                sheet.cell(row = 2+i+accumulate_sum, column = 13).value = Player_Pitcher_Stats[i]['H']
                sheet.cell(row = 2+i+accumulate_sum, column = 14).value = Player_Pitcher_Stats[i]['HR']
                sheet.cell(row = 2+i+accumulate_sum, column = 15).value = Player_Pitcher_Stats[i]['SAC']
                sheet.cell(row = 2+i+accumulate_sum, column = 16).value = Player_Pitcher_Stats[i]['SF']
                sheet.cell(row = 2+i+accumulate_sum, column = 17).value = Player_Pitcher_Stats[i]['BB']
                sheet.cell(row = 2+i+accumulate_sum, column = 18).value = Player_Pitcher_Stats[i]['HBP']
                sheet.cell(row = 2+i+accumulate_sum, column = 19).value = Player_Pitcher_Stats[i]['K']
                sheet.cell(row = 2+i+accumulate_sum, column = 20).value = Player_Pitcher_Stats[i]['WP']
                sheet.cell(row = 2+i+accumulate_sum, column = 21).value = Player_Pitcher_Stats[i]['BK']
                sheet.cell(row = 2+i+accumulate_sum, column = 22).value = Player_Pitcher_Stats[i]['R']
                sheet.cell(row = 2+i+accumulate_sum, column = 23).value = Player_Pitcher_Stats[i]['ER']
                sheet.cell(row = 2+i+accumulate_sum, column = 24).value = Player_Pitcher_Stats[i]['PI']
                if(Player_Pitcher_Stats[i]['IP'] != 0):
                        Player_Pitcher_Stats[i]['ERA'] = (Player_Pitcher_Stats[i]['ER'] * 9) / (Player_Pitcher_Stats[i]['IP'] / 3)
                sheet.cell(row = 2+i+accumulate_sum, column = 25).value = Player_Pitcher_Stats[i]['ERA']
#                sheet.cell(row = 2+i+accumulate_sum, column = 26).value = Player_Pitcher_Stats[i]['FIP']
#                sheet.cell(row = 2+i+accumulate_sum, column = 27).value = Player_Pitcher_Stats[i]['OPS']
#                sheet.cell(row = 2+i+accumulate_sum, column = 28).value = Player_Pitcher_Stats[i]['WAR']
#                sheet.cell(row = 2+i+accumulate_sum, column = 29).value = Player_Pitcher_Stats[i]['GO']
#                sheet.cell(row = 2+i+accumulate_sum, column = 30).value = Player_Pitcher_Stats[i]['FO']
#                sheet.cell(row = 2+i+accumulate_sum, column = 31).value = Player_Pitcher_Stats[i]['LO']
#                sheet.cell(row = 2+i+accumulate_sum, column = 32).value = Player_Pitcher_Stats[i]['P']
#                sheet.cell(row = 2+i+accumulate_sum, column = 33).value = Player_Pitcher_Stats[i]['C']
#                sheet.cell(row = 2+i+accumulate_sum, column = 34).value = Player_Pitcher_Stats[i]['1B']
#                sheet.cell(row = 2+i+accumulate_sum, column = 35).value = Player_Pitcher_Stats[i]['2B']
#                sheet.cell(row = 2+i+accumulate_sum, column = 36).value = Player_Pitcher_Stats[i]['3B']
#                sheet.cell(row = 2+i+accumulate_sum, column = 37).value = Player_Pitcher_Stats[i]['SS']
#                sheet.cell(row = 2+i+accumulate_sum, column = 38).value = Player_Pitcher_Stats[i]['LF']
#                sheet.cell(row = 2+i+accumulate_sum, column = 39).value = Player_Pitcher_Stats[i]['LC']
#                sheet.cell(row = 2+i+accumulate_sum, column = 40).value = Player_Pitcher_Stats[i]['CF']
#                sheet.cell(row = 2+i+accumulate_sum, column = 41).value = Player_Pitcher_Stats[i]['RC']
#                sheet.cell(row = 2+i+accumulate_sum, column = 42).value = Player_Pitcher_Stats[i]['RF']

        accumulate_sum += len(name)        
        print(team_name[0].text + " (" + club_idt[teams] + ") 기록 완료!!!!!")
        #엑셀 파일 만들기
        excel_file.save(path2 + '[투구분석] '+ year + ' AUBL.xlsx')
        #excel_file.save(path + '[일괄분석] '+ year + ' ' + team_name[0].text +'.xlsx')
        #excel_file.close()
        
#excel_file.save(path2 + '[분석] '+ year + ' AUBL.xlsx')
excel_file.close()
        

        

