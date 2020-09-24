import bs4
import urllib.request
import openpyxl
import os
from openpyxl import load_workbook

#playerid bug
#연도 입력
year = str(2019)
club_idt = [0 for i in range(40)]
club_idt[0] = '32122'#한체대
club_idt[1] = '33835'#가천대
club_idt[2] = '23604'#가톨릭대
club_idt[3] = '11841'#건대글로컬
club_idt[4] = '11866'#건대서울
club_idt[5] = '24037'#경기대
club_idt[6] = '23749'#경희국제
club_idt[7] = '23553'#경희서울
club_idt[8] = '11853'#고려대
club_idt[9] = '23999'#광운대
club_idt[10] = '11714'#국민대
club_idt[11] = '21584'#단국죽전
club_idt[12] = '11869'#단국천안
club_idt[13] = '23987'#동국대
club_idt[14] = '11849'#명지서울
club_idt[15] = '23995'#명지용인
club_idt[16] = '25226'#백석대
club_idt[17] = '23515'#상명대
club_idt[18] = '11721'#서강대
club_idt[19] = '23686'#서경대
club_idt[20] = '970'#서울과학기술대
club_idt[21] = '11709'#서울시립대
club_idt[22] = '24158'#성균관대
club_idt[23] = '11810'#세종대
club_idt[24] = '11877'#숭실대
club_idt[25] = '343'#아주대
club_idt[26] = '11919'#연세대
club_idt[27] = '11847'#외대글로벌
club_idt[28] = '23676'#인천대
club_idt[29] = '471'#인하대
club_idt[30] = '24016'#중앙대
club_idt[31] = '28867'#한국교통대
club_idt[32] = '6205'#산기대
club_idt[33] = '23728'#한국외대서울
club_idt[34] = '23593'#항공대
club_idt[35] = '24010'#한성대
club_idt[36] = '23691'#한신대
club_idt[37] = '24012'#한양대에리카
club_idt[38] = '23961'#한양대서울
club_idt[39] = '23511'#홍익대

accumulate_sum=0

#엑셀 파일 관련
#import openpyxl
excel_file = openpyxl.Workbook()
excel_file.remove(excel_file.active)
excel_sheet = excel_file.create_sheet('선수 이름')
excel_sheet = excel_file.create_sheet('팀 기록1') # 조회시 에러 방지
sheet = excel_file.active
       
for teams in range(0, len(club_idt)):
        club_id = club_idt[teams]
        url = "http://www.gameone.kr/club/info/player?club_idx="+ club_id# 한체대 주소
        #향후 32122 이 부분을 한체대에 배정할 예정 ex) 한체대 = 32122
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")

        #팀 이름
        team_name = bsObj.select('title') # team_name[0].text 로 사용


        #선수 이름 주소
        player_name = bsObj.select('form > dt')

        #선수 이름 양식
        player_array = [0 for i in range(len(player_name))]

        count=0
        for num in range(0, len(player_name)):
                for back_num in range(0,1):
                        player_array[count] = player_name[num].text.split('.')
                        count+=1
        #이름만 저장
        name = [0 for i in range(len(player_name))]
        for num in range(0, len(player_array)):
                name[num] = player_array[num][1]

        #등번호만 저장
        backnum = [0 for i in range(len(player_name))]
        for num in range(0, len(player_array)):
                backnum[num] = player_array[num][0]

        #기록 형식에 맞게 이름 저장 윤주엽(5)
        #player = [0 for i in range(len(player_name))]
        #for num in range(0, len(player_array)):
        #        player[num] = name[num] + '(' + backnum[num] + ')'


        #여러 엑셀 파일 읽기
        #import os
        #경로 지정
        #path = "/Users/YEO/AppData/Local/Programs/Python/Python38/"+ year + "/" + team_name[0].text + "/"
        path = "/Users/JuYeop/AppData/Local/Programs/Python/Python38-32/"+ year + "/" + team_name[0].text + "/"
        file_list = os.listdir(path)
        
        ###타자 기록을 위해 변수 선언
        Player_Batter_Stats = [0 for i in range(len(name))]

        for i in range(0, len(name)):
                Player_Batter_Stats[i] = dict.fromkeys({'playerId', 'year', 'G', 'PA', 'AB', 'Runs', 'Hits', '_2B', '_3B',
                                'HR', 'TB', 'RBI', 'SB', 'CS', 'BB', 'IBB', 'HBP', 'SO', 'GDP',
                                'SAC', 'SF', 'AVG', 'OBP', 'SLG', 'OPS', 'WAR', 'GO', 'FO', 'LO', ##GroundBall, Flyball, Linedrive
                                'P', 'C', '1B', '2B', '3B', 'SS', 'LF', 'LC', 'CF', 'RC', 'RF'},0)
        
        if(os.path.exists(path) != True): # 기록 없으면 continue
                print(team_name[0].text + " (" + club_id + ") 기록 없음")
                continue
        

        #폴더 조회 후 없으면 만들기
        if(os.path.exists(path) != True): # 폴더가 존재하지 않는다면
                os.makedirs(path) # 폴더를 만든다.

        #누적 기록 저장소
        accumulate_record = [0 for i in range(len(name))]
        for i in range(0, len(name)):
                accumulate_record[i] = []


        ###기록 가져오기#
        #from openpyxl import load_workbook

        for repeat in range(0, len(file_list)):
                load_wb = load_workbook(path + file_list[repeat], data_only=True)
                load_ws = load_wb['팀 기록1']

                #선수 기록 저장
                player_record = [0 for i in range(len(name))]
                for i in range(0, len(name)):
                        player_record[i] = []
                        
                ################
                #홈 타자 수 + 원정 타자 수 + 1, '이닝' 나올때까지 찾는 원리
                batter_number=0

                for i in range(0, len(name)):
                        if(load_ws.cell(7+i,3).value != '이닝'):
                                batter_number+=1
                        else:
                                break
                        
                for j in range(0, len(name)): #전체선수숫자조회
                        for i in range(0, batter_number): # 원정타자수 + 홈타자수 + 1로 변경해야#############
                                if(load_ws.cell(7+i,3).value != None):
                                        if(name[j] == load_ws.cell(7+i,3).value.split('(')[0]): # 게임 기록에서 선수이름 조회
                                                for k in range(1, 13): # 12이닝
                                                        if(load_ws.cell(7+i,3+k).value != None): # 기록이 비어있지 않다면
                                                                player_record[j].append(load_ws.cell(7+i,3+k).value) # 선수기록항목에 저장

                #도루, 사구 등 세부기록 분리를 위해 배열을 문자열로 만든 뒤, 문자열을 배열로 만든다.
                # , 추가
                values = [0 for i in range(len(name))]

                for i in range(0, len(name)):
                        values[i] = ",".join(str(v) for v in player_record[i]).replace("/",",") #,를 기준으로 리스트에서 문자열로 변경, /를 ,로 변경


                #문자열을 배열로
                record_split = [0 for i in range(len(name))]

                #, 기준으로 나눈다
                for i in range(0, len(name)):
                        record_split[i] = values[i].split(",")


                #누적 기록 저장
                for i in range(0, len(name)):
                        accumulate_record[i] += record_split[i]

                ################
                #게임수 넣기
                for i in range(0, batter_number): #전체선수숫자 조회
                        if(load_ws.cell(7+i,3).value != None):
                                for j in range(0, len(name)): # 원정타자수 + 홈타자수 + 1 로 변경해야########
                                        if(load_ws.cell(7+i,3).value.split('(')[0] == name[j]):
                                                Player_Batter_Stats[j]['G'] +=1

                ################
                #타점 기록하기
                for i in range(0, batter_number): #선수 조회
                        if(load_ws.cell(7+i,18).value != None):
                                for j in range(0, len(name)):
                                        if(load_ws.cell(7+i,3).value.split('(')[0] == name[j]):
                                                Player_Batter_Stats[j]['RBI'] += int(load_ws.cell(7+i,18).value)
                                                Player_Batter_Stats[j]['Runs'] += int(load_ws.cell(7+i,19).value) #득점
                                                break


        #빈 배열 삭제
        for i in range(0, len(name)):
                accumulate_record[i] = [v for v in accumulate_record[i] if v]

        #기록이 잘 넘어가는지 확인하기 위해
        #기록을 넘기면 rowdata를 지우며 진행하기 위해
        #backup 변수 선언
        backup = accumulate_record


        hi=0 # 총 자료 개수 조회
        for i in range(len(backup)):
                hi += len(backup[i])
                
        count=0 #몇 개 조회되었나 확인중.


        for i in range(0, len(name)): #선수 조회
                for j in range(0, len(backup[i])): # 전체 기록 조회
        #삼진
                        if(backup[i][j] == '삼진'): # 삼진, code 0
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['SO'] +=1
                                count+=1
        #4구
                        elif(backup[i][j] == '4구'): # 4구, code 1
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['BB'] +=1
                                count+=1
        #사구
                        elif(backup[i][j] == '사구'): # 사구, code 2
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['HBP'] +=1
                                count+=1
        #낫아웃-
                        elif(backup[i][j].find("낫아웃-") != -1): # 낫아웃-, code 20
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['SO'] +=1
                                count+=1
        #낫아웃+                        
                        elif(backup[i][j].find('낫아웃+') != -1): # 낫아웃출루, code 21
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['SO'] +=1
                                count+=1
        #타격방해
                        elif(backup[i][j] == '타격방해'): # 타격방해, code 23
                                Player_Batter_Stats[i]['PA'] +=1
                                count+=1
        #도루
                        elif(backup[i][j] == '도루'): # 도루, code 
                                Player_Batter_Stats[i]['SB'] +=1           
                                count+=1
        #도루자
                        elif(backup[i][j].find('도루자') != -1): # 도루자, code 
                                Player_Batter_Stats[i]['CS'] +=1           
                                count+=1
        #주자아웃
                        elif(backup[i][j].find('주자아웃') != -1): # 주루사, code 
                                count+=1
        #주자아웃
                        elif(backup[i][j].find('주루사') != -1): # 주루사, code 
                                count+=1
        #폭투                        
                        elif(backup[i][j] == '폭투'): # 폭투, code 
                                count+=1
        #포일
                        elif(backup[i][j] == '포일'): # 포일, code 
                                count+=1
        #보크
                        elif(backup[i][j] == '보크'): # 보크, code 
                                count+=1                        
        #송구실책
                        elif(backup[i][j].find("송구실책") != -1): # 송구실책, code 
                                count+=1 
        #포구실책
                        elif(backup[i][j].find("포구실책") != -1): # 포구실책, code 
                                count+=1 
        #주루방해
                        elif(backup[i][j] == '주루방해'): # 주루방해, code 
                                count+=1 
        #대수비
                        elif(backup[i][j] == '대수비'): # 대수비, code 
                                count+=1 
        #대주자
                        elif(backup[i][j] == '대주자'): # 대주자, code 
                                count+=1 
        #대타
                        elif(backup[i][j] == '대타'): # 대타, code 
                                count+=1 
        #승부주자
                        elif(backup[i][j] == '승부주자'): # 승부주자, code 
                                count+=1 
        #타구맞음
                        elif(backup[i][j] == '타구맞음'): # 타구맞음, code 
                                count+=1 
        #런다운
                        elif(backup[i][j] == '런다운'): # 런다운, code 
                                count+=1 
        #수비방해
                        elif(backup[i][j] == '수비방해'): # 수비방해, code 
                                count+=1 
        #고의사구
                        elif(backup[i][j] == '고의4구'): # 고의4구, code 
                                Player_Batter_Stats[i]['IBB'] +=1
                                count+=1
        #쓰리번트
                        elif(backup[i][j] == '쓰리번트'): # 쓰리번트, code 
                                Player_Batter_Stats[i]['SO'] +=1
                                count+=1
        #타자타구
                        elif(backup[i][j] == '타자타구'): # 타구맞아아웃, code 
                                count+=1                        
        #부정타격
                        elif(backup[i][j] == '부정타격'): #부정타격아웃, code 
                                count+=1
        #견제사
                        elif(backup[i][j].find("견제사") != -1): #견제사, code 
                                count+=1
        #런다운
                        elif(backup[i][j] == '런다운'): #견제사, code 
                                count+=1
        #땅볼R
                        elif(backup[i][j] == '투땅R'): # 투수 땅볼R, code 41
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포땅R'): # 포수 땅볼R, code 42
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1땅R'): # 1루수 땅볼R, code 43
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2땅R'): # 2루수 땅볼R, code 44
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3땅R'): # 3루수 땅볼R, code 45
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유땅R'): # 유격수 땅볼R, code 46
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌땅R'): # 좌익수 땅볼R, code 47
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중땅R'): # 중견수 땅볼R, code 48
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우땅R'): # 우익수 땅볼R, code 49
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #땅볼
                        elif(backup[i][j].find('투땅') != -1): # 투수 땅볼, code 31
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1 #GroundOut
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j].find('포땅') != -1): # 포수 땅볼, code 32
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j].find('1땅') != -1): # 1루수 땅볼, code 33
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j].find('2땅') != -1): # 2루수 땅볼, code 34
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find('3땅') != -1): # 3루수 땅볼, code 35
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j].find('유땅') != -1): # 유격수 땅볼, code 36
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j].find('좌땅') != -1): # 좌익수 땅볼, code 37
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j].find('중땅') != -1): # 중견수 땅볼, code 38
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j].find('우땅') != -1): # 우익수 땅볼, code 39
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #뜬공
                        elif(backup[i][j] == '투플'): # 투수 플라이, code 51
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포플'): # 포수 플라이, code 52
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1플'): # 1루수 플라이, code 53
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2플'): # 2루수 플라이, code 54
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3플'): # 3루수 플라이, code 55
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유플'): # 유격수 플라이, code 56
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌플'): # 좌익수 플라이, code 57
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중플'): # 중견수 플라이, code 58
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우플'): # 우익수 플라이, code 59
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #직선타
                        elif(backup[i][j] == '투직'): # 투수 직선타, code 61
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1 # Linedrive Out
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포직'): # 포수 직선타, code 62
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1직'): # 1루수 직선타, code 63
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2직'): # 2루수 직선타, code 64
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3직'): # 3루수 직선타, code 65
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유직'): # 유격수 직선타, code 66
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌직'): # 좌익수 직선타, code 67
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중직'): # 중견수 직선타, code 68
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우직'): # 우익수 직선타, code 69
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #희생플라이
                        elif(backup[i][j] == '투희플'): # 투수 희생플라이, code 71
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['SF'] +=1 # Sacrifice Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포희플'): # 포수 희생플라이, code 72
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1희플'): # 1루수 희생플라이, code 73
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2희플'): # 2루수 희생플라이, code 74
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3희플'): # 3루수 희생플라이, code 75
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유희플'): # 유격수 희생플라이, code 76
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '투희플'): # 투수 파울희생플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['SF'] +=1 # Sacrifice Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포파희플'): # 포수 파울희생플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1파희플'): # 1루수 파울희생플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2파희플'): # 2루수 파울희생플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3파희플'): # 3루수 파울희생플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유파희플'): # 유격수 파울희생플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '좌희플'): # 좌익수 희생플라이, code 77
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중희플'): # 중견수 희생플라이, code 78
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우희플'): # 우익수 희생플라이, code 79
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
                        elif(backup[i][j] == '희생플'): # 희생플라이, code
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                count+=1

                        elif(backup[i][j] == '좌파희플'): # 좌익수 파울희생플라이, code 77
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '우파희플'): # 우익수 파울희생플라이, code 79
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #희플출루
                        elif(backup[i][j] == '투희플출'): # 투수 희생플라이출루, code
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['SF'] +=1 # Sacrifice Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포희플출'): # 포수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1희플출'): # 1루수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2희플출'): # 2루수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3희플출'): # 3루수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유희플출'): # 유격수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌희플출'): # 좌익수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중희플출'): # 중견수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우희플출'): # 우익수 희생플라이출루, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SF'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #희생번트
                        elif(backup[i][j] == '투희번'): # 투수 희생번트, code 81
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포희번'): # 포수 희생번트, code 82
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['C'] +=1               
                                count+=1

                        elif(backup[i][j] == '1희번'): # 1루수 희생번트, code 83
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['1B'] +=1               
                                count+=1

                        elif(backup[i][j] == '2희번'): # 2루수 희생번트, code 84
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['2B'] +=1               
                                count+=1

                        elif(backup[i][j] == '3희번'): # 3루수 희생번트, code 85
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['3B'] +=1               
                                count+=1

                        elif(backup[i][j] == '유희번'): # 유격수 희생번트, code 86
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['SS'] +=1               
                                count+=1

                        elif(backup[i][j] == '좌희번'): # 좌익수 희생번트, code 87
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['LF'] +=1               
                                count+=1

                        elif(backup[i][j] == '중희번'): # 중견수 희생번트, code 88
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['CF'] +=1               
                                count+=1

                        elif(backup[i][j] == '우희번'): # 우익수 희생번트, code 89
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['SAC'] +=1
                                Player_Batter_Stats[i]['RF'] +=1               
                                count+=1
        #병살타
                        elif(backup[i][j] == '투땅병살'): # 투수 병살타, code 91
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1 #GroundOut
                                Player_Batter_Stats[i]['GDP'] +=1 #Ground into Double Play
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포땅병살'): # 포수 병살타, code 92
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j].find('1땅병살') != -1): # 1루수 병살타, code 93
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j].find('2땅병살') != -1): # 2루수 병살타, code 94
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find('3땅병살') != -1): # 3루수 병살타, code 95
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j].find('유땅병살') != -1): # 유격수 병살타, code 96
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌땅병살'): # 좌익수 병살타, code 97
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중땅병살'): # 중견수 병살타, code 98
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우땅병살'): # 우익수 병살타, code 99
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #직선타병살
                        elif(backup[i][j] == '투직병살'): # 투수 직선병살타, code 101
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1 # Linedrive Out
                                Player_Batter_Stats[i]['GDP'] +=1 #Ground into Double Play
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포직병살'): # 포수 직선병살타, code 102
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1직병살'): # 1루수 직선병살타, code 103
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2직병살'): # 2루수 직선병살타, code 104
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3직병살'): # 3루수 직선병살타, code 105
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유직병살'): # 유격수 직선병살타, code 106
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌직병살'): # 좌익수 직선병살타, code 107
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중직병살'): # 중견수 직선병살타, code 108
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우직병살'): # 우익수 직선병살타, code 109
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #뜬공병살
                        elif(backup[i][j] == '투플병살'): # 투수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 #FlyOut
                                Player_Batter_Stats[i]['GDP'] +=1 #Ground into Double Play
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포플병살'): # 포수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1플병살'): # 1루수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2플병살'): # 2루수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3플병살'): # 3루수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유플병살'): # 유격수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌플병살'): # 좌익수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중플병살'): # 중견수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우플병살'): # 우익수 뜬공병살타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #땅볼 삼중살
        #게임원엔 없음
                        elif(backup[i][j] == '투땅삼중살'): # 투수 땅볼삼중살, code 111
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1 #GroundOut
                                Player_Batter_Stats[i]['GDP'] +=1 #Ground into Double Play
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포땅삼중살'): # 포수 땅볼삼중살, code 112
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1땅삼중살'): # 1루수 땅볼삼중살, code 113
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2땅삼중살'): # 2루수 땅볼삼중살, code 114
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3땅삼중살'): # 3루수 땅볼삼중살, code 115
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유땅삼중살'): # 유격수 땅볼삼중살, code 116
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌땅삼중살'): # 좌익수 땅볼삼중살, code 117
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중땅삼중살'): # 중견수 땅볼삼중살, code 118
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우땅삼중살'): # 우익수 땅볼삼중살, code 119
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1        
        #직선타 삼중살
        #게임원엔 없음
                        elif(backup[i][j] == '투직삼중살'): # 투수 직선삼중살, code 111
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1 # Linedrive Out
                                Player_Batter_Stats[i]['GDP'] +=1 #Ground into Double Play
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포직삼중살'): # 포수 직선삼중살, code 112
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1직삼중살'): # 1루수 직선삼중살, code 113
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2직삼중살'): # 2루수 직선삼중살, code 114
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3직삼중살'): # 3루수 직선삼중살, code 115
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유직삼중살'): # 유격수 직선삼중살, code 116
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌직삼중살'): # 좌익수 직선삼중살, code 117
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중직삼중살'): # 중견수 직선삼중살, code 118
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우직삼중살'): # 우익수 직선삼중살, code 119
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #안타
                        elif(backup[i][j] == '투안'): # 투수 안타, code 121
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포안'): # 포수 안타, code 122
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1안'): # 1루수 안타, code 123
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2안'): # 2루수 안타, code 124
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3안'): # 3루수 안타, code 125
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유안'): # 유격수 안타, code 126
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌안'): # 좌익수 안타, code 127
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '좌전안'): # 좌익수 안타, code 127
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '좌선안'): # 좌익수 안타, code 127
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '좌중안'): # 좌중간 안타, code 1220
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['LC'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '좌월안'): # 좌익수 안타, code 1274
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '중안'): # 중견수 안타, code 128
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '중월안'): # 중견수 안타, code 128
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '중전안'): # 중견수 안타, code 128
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '우중안'): # 우중간 안타, code 1210
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['RC'] +=1
                                count+=1

                        elif(backup[i][j] == '우안'): # 우익수 안타, code 1210
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1                        

                        elif(backup[i][j] == '우전안'): # 우익수 안타, code 1211
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1  

                        elif(backup[i][j] == '우선안'): # 우익수 안타, code 1211
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '우월안'): # 우익수 안타, code 1214
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #내야안타
                        elif(backup[i][j] == '투내안'): # 투수 내야안타, code 131
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포내안'): # 포수 내야안타, code 132
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1내안'): # 1루수 내야안타, code 133
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2내안'): # 2루수 내야안타, code 134
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3내안'): # 3루수 내야안타, code 135
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유내안'): # 유격수 내야안타, code 136
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1    
        #2루타
                        elif(backup[i][j] == '투2'): # 투수 2루타, code 141
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['_2B'] +=1 # 2Hit
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포2'): # 포수 2루타, code 142
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '12'): # 1루수 2루타, code 143
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '22'): # 2루수 2루타, code 144
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '32'): # 3루수 2루타, code 145
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유2'): # 유격수 2루타, code 146
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌2'): # 좌익수 2루타, code 147
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌안2'): # 좌익수 2루타, code 147
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '중2'): # 중견수 2루타, code 148
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                        elif(backup[i][j] == '중안2'): # 중견수 2루타, code 148
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '우2'): # 우익수 2루타, code 149
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
                        elif(backup[i][j] == '우안2'): # 우익수 2루타, code 149
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #외야 2루타 모음                        
                        elif(backup[i][j] == '좌선2'): # 좌익선상 2루타, code 1472
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌전2'): # 좌익수 앞 2루타, code 1471
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌월2'): # 좌익수 뒤 2루타, code 1474
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌중2'): # 좌중간 2루타, code 1420
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LC'] +=1
                                count+=1
                                       
                        elif(backup[i][j] == '중전2'): # 중견수 앞 2루타, code 1481
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1                       
                        elif(backup[i][j] == '중월2'): # 중견수 뒤 2루타, code 1484
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                        elif(backup[i][j] == '중선2'): # 중견수 2루타, code 148
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '우중2'): # 우중간 2루타, code 1410
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RC'] +=1
                                count+=1
                        elif(backup[i][j] == '우전2'): # 우익수 앞 2루타, code 1491
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우월2'): # 우익수 뒤 2루타, code 1494
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우선2'): # 우익선상 2루타, code 1493
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
        #3루타                        
                        elif(backup[i][j] == '투3'): # 투수 3루타, code 151
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['_3B'] +=1 # 3Hit
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포3'): # 포수 3루타, code 152
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '13'): # 1루수 3루타, code 153
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '23'): # 2루수 3루타, code 154
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '33'): # 3루수 3루타, code 155
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유3'): # 유격수 3루타, code 156
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌3'): # 좌익수 3루타, code 157
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '좌안3'): # 좌익수 3루타, code 157
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '중3'): # 중견수 3루타, code 158
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '중안3'): # 중견수 3루타, code 158
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '우3'): # 우익수 3루타, code 159
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '우안3'): # 우익수 3루타, code 159
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #외야 3루타 모음                        
                        elif(backup[i][j] == '좌선3'): # 좌익선상 3루타, code 1572
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌전3'): # 좌익수 앞 3루타, code 1571
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌월3'): # 좌익수 뒤 3루타, code 1574
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌중3'): # 좌중간 3루타, code 1520
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['LC'] +=1
                                count+=1
                                       
                        elif(backup[i][j] == '중전3'): # 중견수 앞 3루타, code 1581
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1                       
                        elif(backup[i][j] == '중월3'): # 중견수 뒤 3루타, code 1584
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                               
                        elif(backup[i][j] == '우중3'): # 우중간 3루타, code 1510
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['RC'] +=1
                                count+=1
                        elif(backup[i][j] == '우전3'): # 우익수 앞 3루타, code 1591
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우월3'): # 우익수 뒤 3루타, code 1594
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우선3'): # 우익선상 3루타, code 1593
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #G홈
                        elif(backup[i][j] == '투G홈'): # 투수 G홈, code 161
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['HR'] +=1 # HR
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포G홈'): # 포수 G홈, code 162
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1G홈'): # 1루수 G홈, code 163
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2G홈'): # 2루수 G홈, code 164
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3G홈'): # 3루수 G홈, code 165
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유G홈'): # 유격수 G홈, code 166
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌G홈'): # 좌익수 G홈, code 167
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중G홈'): # 중견수 G홈, code 168
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우G홈'): # 우익수 G홈, code 169
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #외야 G홈 모음                        
                        elif(backup[i][j] == '좌선G홈'): # 좌익선상 G홈, code 1672
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌전G홈'): # 좌익수 앞 G홈, code 1671
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌월G홈'): # 좌익수 뒤 G홈, code 1674
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌중G홈'): # 좌중간 G홈, code 1620
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LC'] +=1
                                count+=1
                                       
                        elif(backup[i][j] == '중전G홈'): # 중견수 앞 G홈, code 1681
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1                       
                        elif(backup[i][j] == '중월G홈'): # 중견수 뒤 G홈, code 1684
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                               
                        elif(backup[i][j] == '우중G홈'): # 우중간 G홈, code 1610
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RC'] +=1
                                count+=1
                        elif(backup[i][j] == '우전G홈'): # 우익수 앞 G홈, code 1691
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우월G홈'): # 우익수 뒤 G홈, code 1694
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우선G홈'): # 우익선상 G홈, code 1693
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #홈런

                        elif(backup[i][j] == '좌홈'): # 좌익수 홈런, code 177
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중홈'): # 중견수 홈런, code 178
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우홈'): # 우익수 홈런, code 179
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #외야 홈 모음                        
                        elif(backup[i][j] == '좌선홈'): # 좌익선상 홈런, code 1772
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌월홈'): # 좌익수 뒤 홈런, code 1774
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '좌중홈'): # 좌중간 홈런, code 1720
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['LC'] +=1
                                count+=1
                                       
                        elif(backup[i][j] == '중월홈'): # 중견수 뒤 홈런, code 1784
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                               
                        elif(backup[i][j] == '우중홈'): # 우중간 홈런, code 1710
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RC'] +=1
                                count+=1
                        elif(backup[i][j] == '우월홈'): # 우익수 뒤 홈런, code 1794
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '우선홈'): # 우익선상 홈런, code 1793
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['HR'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #번땅, 번트도 땅볼아웃에 포함해서 계산하나
                        elif(backup[i][j].find('투번') != -1): # 투수 번트땅볼, code 181
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1 #GroundOut
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j].find('포번') != -1): # 포수 번트땅볼, code 182
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j].find('1번') != -1): # 1루수 번트땅볼, code 183
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j].find('2번') != -1): # 2루수 번트땅볼, code 184
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find('3번') != -1): # 3루수 번트땅볼, code 185
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j].find('유번') != -1): # 유격수 번트땅볼, code 186
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌번'): # 좌익수 번트땅볼, code 187
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중번'): # 중견수 번트땅볼, code 188
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우번'): # 우익수 번트땅볼, code 189
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #번뜬, 번트도 뜬공아웃에 포함해서 계산하나
                        elif(backup[i][j] == '투번플'): # 투수 번트플라이, code 191
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포번플'): # 포수 번트플라이, code 192
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1번플'): # 1루수 번트플라이, code 193
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2번플'): # 2루수 번트플라이, code 194
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3번플'): # 3루수 번트플라이, code 195
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유번플'): # 유격수 번트플라이, code 196
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌번플'): # 좌익수 번트플라이, code 197
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중번플'): # 중견수 번트플라이, code 198
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우번플'): # 우익수 번트플라이, code 199
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #야수선택, 땅볼로 간주
                        elif(backup[i][j].find("투야선") != -1): # 투수 야수선택, code 221
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1 #GroundOut
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j].find("포야선") != -1): # 포수 야수선택, code 222
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j].find("1야선") != -1): # 1루수 야수선택, code 223
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j].find("2야선") != -1): # 2루수 야수선택, code 224
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find("3야선") != -1): # 3루수 야수선택, code 225
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j].find("유야선") != -1): # 유격수 야수선택, code 226
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j].find("좌야선") != -1): # 좌익수 야수선택, code 227
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j].find("중야선") != -1): # 중견수 야수선택, code 228
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j].find("우야선") != -1): # 우익수 야수선택, code 229
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #인필드플라이
                        elif(backup[i][j] == '투인플'): # 투수 인필드플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포인플'): # 포수 인필드플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1인플'): # 1루수 인필드플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2인플'): # 2루수 인필드플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3인플'): # 3루수 인필드플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유인플'): # 유격수 인필드플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1
        #실책
                        elif(backup[i][j].find('투실') != -1): # 투수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j].find('포실') != -1): # 포수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j].find('1실') != -1): # 1루수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j].find('2실') != -1): # 2루수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find('3실') != -1): # 3루수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j].find('유실') != -1): # 유격수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌실'): # 좌익수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중실'): # 중견수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우실'): # 우익수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1                        

                        elif(backup[i][j].find('투땅실') != -1): # 투수 실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j].find('포땅실') != -1): # 포수 땅볼실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j].find('1땅실') != -1): # 1루수 땅볼실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j].find('2땅실') != -1): # 2루수 땅볼실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find('3땅실') != -1): # 3루수 땅볼실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j].find('유땅실') != -1): # 유격수 땅볼실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '포플실'): # 포수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1플실'): # 1루수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2플실'): # 2루수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3플실'): # 3루수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유플실'): # 유격수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '좌플실'): # 좌익수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중플실'): # 중견수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우플실'): # 우익수 플라이실책, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1                                
        #파울플라이
                        elif(backup[i][j] == '투파플'): # 투수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 # Flyout
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포파플'): # 포수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1파플'): # 1루수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2파플'): # 2루수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3파플'): # 3루수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유파플'): # 유격수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1

                        elif(backup[i][j] == '좌파플'): # 좌익수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1

                        elif(backup[i][j] == '중파플'): # 중견수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1

                        elif(backup[i][j] == '우파플'): # 우익수 파울플라이, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1 
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
        #번트안타
                        elif(backup[i][j] == '투번안'): # 투수 번트안타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['P'] +=1               
                                count+=1

                        elif(backup[i][j] == '포번안'): # 포수 번트안타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['C'] +=1
                                count+=1

                        elif(backup[i][j] == '1번안'): # 1루수 번트안타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['1B'] +=1
                                count+=1

                        elif(backup[i][j] == '2번안'): # 2루수 번트안타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j] == '3번안'): # 3루수 번트안타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['3B'] +=1
                                count+=1

                        elif(backup[i][j] == '유번안'): # 유격수 번트안타, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['SS'] +=1
                                count+=1
        #인정2루타
                        elif(backup[i][j] == '인좌선2'): # 좌익선상 2루타, code 1472
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '인좌월2'): # 좌익수 뒤 2루타, code 1474
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                        elif(backup[i][j] == '인좌중2'): # 좌중간 2루타, code 1420
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LC'] +=1
                                count+=1
                        elif(backup[i][j] == '인좌2'): # 좌익수 2루타, code 147
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['LF'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '인중월2'): # 중견수 뒤 2루타, code 1484
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['CF'] +=1
                                count+=1
                               
                        elif(backup[i][j] == '인우중2'): # 우중간 2루타, code 1410
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RC'] +=1
                                count+=1
                        elif(backup[i][j] == '인우월2'): # 우익수 뒤 2루타, code 1494
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1   
                        elif(backup[i][j] == '인우선2'): # 우익선상 2루타, code 1493
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1
                        elif(backup[i][j] == '인우2'): # 우익수 2루타, code 149
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['RF'] +=1
                                count+=1                        
        #낫아웃                        
                        elif(backup[i][j] == '낫아웃'): # 낫아웃출루, code 21
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['SO'] +=1
                                count+=1                        
        #도루사
                        elif(backup[i][j] == '도루사'): # 도루사, code 
                                Player_Batter_Stats[i]['CS'] +=1           
                                count+=1
        #1루타
                        elif(backup[i][j] == '1루타'): # 1루타
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                count+=1
        #2루타
                        elif(backup[i][j] == '2루타'): # 2루타
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                count+=1
        #3루타
                        elif(backup[i][j] == '3루타'): # 3루타
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['_3B'] +=1
                                count+=1
        #홈런
                        elif(backup[i][j] == '홈런'): # 홈런
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['HR'] +=1
                                count+=1
                                
                        elif(backup[i][j] == '병살'): # 병살
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GDP'] +=1
                                count+=1

                        elif(backup[i][j] == '실책'): # 실책
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                count+=1

                        elif(backup[i][j].find('출루') != -1): # 출루
                                count+=1                                
                        elif(backup[i][j] == '희생'): # 희생
                                Player_Batter_Stats[i]['PA'] +=1
                                count+=1
                        elif(backup[i][j] == '야선'): # 야선
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1                                
                                count+=1
                        elif(backup[i][j] == '아웃'): # 아웃
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                count+=1                                
                        elif(backup[i][j] == '인정2'): # 인정2
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                count+=1
                        elif(backup[i][j] == '인플'): # 인필드플라이
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                count+=1
                        elif(backup[i][j] == '플라이'): # 플라이
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['FO'] +=1
                                count+=1
                        elif(backup[i][j] == '땅볼'): # 플라이
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['GO'] +=1
                                count+=1
                        elif(backup[i][j] == '직선타'): # 플라이
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['LO'] +=1
                                count+=1                                 
                        elif(backup[i][j].find("E") != -1):
                                 for k in range(1,9):
                                        for l in range(1,9):
                                                if(backup[i][j] == str(k) + 'E' + str(l)):
                                                        count+=1
                                                        break                               
                        elif(backup[i][j] == '2내안2'): # 2내안 2루타, code 1474
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1
                                Player_Batter_Stats[i]['_2B'] +=1
                                Player_Batter_Stats[i]['2B'] +=1
                                count+=1

                        elif(backup[i][j].find('리터치') != -1): # 유격수 땅볼실책, code 
                                count+=1
                        elif(backup[i][j].find('G홈') != -1): # 그라운드 홈런, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['HR'] +=1 # HR
                                count+=1
                        elif(backup[i][j] == '타격방'): #타격방해, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                count+=1
                        elif(backup[i][j] == '4G'): #G홈, code 
                                Player_Batter_Stats[i]['PA'] +=1
                                Player_Batter_Stats[i]['AB'] +=1
                                Player_Batter_Stats[i]['Hits'] +=1 # Hit
                                Player_Batter_Stats[i]['HR'] +=1 # HR
                                count+=1
                        else:           
                                print(backup[i][j])

        #타율 계산
        for i in range(0, len(name)): #선수 조회
                if(Player_Batter_Stats[i]['AB'] != 0):
                        #타율 계산, 소수점 자리수 4자리로 제한
                        Player_Batter_Stats[i]['AVG'] = round((Player_Batter_Stats[i]['Hits'] / Player_Batter_Stats[i]['AB']),4)

                        #출루율 계산, 안타+볼넷+사구 / 타수+볼넷+사구+희플
                        Player_Batter_Stats[i]['OBP'] = round((Player_Batter_Stats[i]['Hits'] + Player_Batter_Stats[i]['BB'] + Player_Batter_Stats[i]['HBP'])
                        / (Player_Batter_Stats[i]['AB'] + Player_Batter_Stats[i]['BB'] + Player_Batter_Stats[i]['HBP'] + Player_Batter_Stats[i]['SF']),4)

                        #총 루타수 계산
                        Player_Batter_Stats[i]['TB'] = round((Player_Batter_Stats[i]['Hits'] + 2*Player_Batter_Stats[i]['_2B'] + 3*Player_Batter_Stats[i]['_3B']
                        +Player_Batter_Stats[i]['HR']),4)
                        
                        #장타율 계산, (1루타*1 + 2루타*2 + 3루타*3 + 홈런*4)/타수
                        Player_Batter_Stats[i]['SLG'] = round((Player_Batter_Stats[i]['TB'] / Player_Batter_Stats[i]['AB']),4)

                        #OPS 계산, 출루율 + 장타율
                        Player_Batter_Stats[i]['OPS'] = Player_Batter_Stats[i]['OBP'] + Player_Batter_Stats[i]['SLG']

        teamId = 0
        #teamId 입력받기
        if(club_id == '32122'): #한체대
                teamId = '1'
        elif(club_id == '33835'): # 가천대
                teamId = '2'
        elif(club_id == '23604'): # 가톨릭대
                teamId = '3'     
        elif(club_id == '11841'): # 건대(글로컬)
                teamId = '4'
        elif(club_id == '11866'): # 건대(서울)
                teamId = '5'
                
        elif(club_id == '24037'): # 경기대
                teamId = '6'
        elif(club_id == '23749'): # 경희대(국제)
                teamId = '7'
        elif(club_id == '23553'): # 경희대(서울)
                teamId = '8'
        elif(club_id == '11853'): # 고려대
                teamId = '9'
        elif(club_id == '23999'): # 광운대
                teamId = '10'
                
        elif(club_id == '11714'): # 국민대
                teamId = '11'
        elif(club_id == '21584'): # 단국대(죽전)
                teamId = '12'
        elif(club_id == '11869'): # 단국대(천안)
                teamId = '13'
        elif(club_id == '23987'): # 동국대
                teamId = '14'
        elif(club_id == '11849'): # 명지대(서울)
                teamId = '15'
                
        elif(club_id == '23995'): # 명지대(용인)
                teamId = '16'
        elif(club_id == '25226'): # 백석대
                teamId = '17'
        elif(club_id == '23515'): # 상명대
                teamId = '18'
        elif(club_id == '11721'): # 서강대
                teamId = '19'
        elif(club_id == '23686'): # 서경대
                teamId = '20'
                
        elif(club_id == '970'): # 서울과학기술대
                teamId = '21'
        elif(club_id == '11709'): # 서울시립대
                teamId = '22'
        elif(club_id == '24158'): # 성균관대
                teamId = '23'
        elif(club_id == '11810'): # 세종대
                teamId = '24'
        elif(club_id == '11877'): # 숭실대
                teamId = '25'
                
        elif(club_id == '343'): # 아주대
                teamId = '26'
        elif(club_id == '11919'): # 연세대
                teamId = '27'
        elif(club_id == '11847'): # 외대(글로벌)
                teamId = '28'
        elif(club_id == '23676'): # 인천대
                teamId = '29'
        elif(club_id == '471'): # 인하대
                teamId = '30'
                
        elif(club_id == '24016'): # 중앙대
                teamId = '31'
        elif(club_id == '28867'): # 한국교통대
                teamId = '32'
        elif(club_id == '6205'): # 한국산기대
                teamId = '33'
        elif(club_id == '23728'): # 한국외대(서울)
                teamId = '34'
        elif(club_id == '23593'): # 한국항공대
                teamId = '35'
                
        elif(club_id == '24010'): # 한성대
                teamId = '36'
        elif(club_id == '23691'): # 한신대
                teamId = '37'
        elif(club_id == '24012'): # 한양대(에리카)
                teamId = '38'
        elif(club_id == '23961'): # 한양대(서울)
                teamId = '39'
        elif(club_id == '23511'): # 홍익대
                teamId = '40'

        #playerId 입력받기
        #path2 = "/Users/YEO/AppData/Local/Programs/Python/Python38/"
        path2 = "/Users/JuYeop/AppData/Local/Programs/Python/Python38-32/"
        file = load_workbook(path2 + "player_info.xlsx", data_only = True)
        file_open = file['Sheet1']

        #행 갯수 세기
        row_count=0
        for row in file_open.rows:
                row_count+=1

        row_count_str = str(row_count)

        for i in range(1, row_count+1):
                if(str(file_open['B'+str(i)].value) == teamId): #학교 탐색
                        for j in range(0, len(name)):
                                if(name[j] == file_open['D'+str(i)].value): # 이름 탐색
                                        Player_Batter_Stats[j]['playerId'] = file_open['A'+str(i)].value
                                        break

        #연도, 이름 양식 입력
        sheet.cell(row = 1, column = 1).value = '선수명'
        sheet.cell(row = 1, column = 2).value = '등번호'
        sheet.cell(row = 1, column = 3).value = 'playerId'
        sheet.cell(row = 1, column = 4).value = 'year'
        sheet.cell(row = 1, column = 5).value = 'G'
        sheet.cell(row = 1, column = 6).value = 'PA'
        sheet.cell(row = 1, column = 7).value = 'AB'
        sheet.cell(row = 1, column = 8).value = 'Runs'
        sheet.cell(row = 1, column = 9).value = 'Hits'
        sheet.cell(row = 1, column = 10).value = '_2B'
        sheet.cell(row = 1, column = 11).value = '_3B'
        sheet.cell(row = 1, column = 12).value = 'HR'
        sheet.cell(row = 1, column = 13).value = 'TB'
        sheet.cell(row = 1, column = 14).value = 'RBI'
        sheet.cell(row = 1, column = 15).value = 'SB'
        sheet.cell(row = 1, column = 16).value = 'CS'
        sheet.cell(row = 1, column = 17).value = 'BB'
        sheet.cell(row = 1, column = 18).value = 'IBB'
        sheet.cell(row = 1, column = 19).value = 'HBP'
        sheet.cell(row = 1, column = 20).value = 'SO'
        sheet.cell(row = 1, column = 21).value = 'GDP'
        sheet.cell(row = 1, column = 22).value = 'SAC'
        sheet.cell(row = 1, column = 23).value = 'SF'
        sheet.cell(row = 1, column = 24).value = 'AVG'
        sheet.cell(row = 1, column = 25).value = 'OBP'
        sheet.cell(row = 1, column = 26).value = 'SLG'
        sheet.cell(row = 1, column = 27).value = 'OPS'
        sheet.cell(row = 1, column = 28).value = 'WAR'
        sheet.cell(row = 1, column = 29).value = 'GO'
        sheet.cell(row = 1, column = 30).value = 'FO'
        sheet.cell(row = 1, column = 31).value = 'LO'
        sheet.cell(row = 1, column = 32).value = 'P'
        sheet.cell(row = 1, column = 33).value = 'C'
        sheet.cell(row = 1, column = 34).value = '1B'
        sheet.cell(row = 1, column = 35).value = '2B'
        sheet.cell(row = 1, column = 36).value = '3B'
        sheet.cell(row = 1, column = 37).value = 'SS'
        sheet.cell(row = 1, column = 38).value = 'LF'
        sheet.cell(row = 1, column = 39).value = 'LC'
        sheet.cell(row = 1, column = 40).value = 'CF'
        sheet.cell(row = 1, column = 41).value = 'RC'
        sheet.cell(row = 1, column = 42).value = 'RF'


        #기록 양식 저장
        num=0
        #Player_Batter_Stats[0].keys()

        #for i in range(0, len(name)):
        #        sheet.cell(row = 1, column = 3+num) = Player_Batter_Stats[i]['playerId']
        #        num+=1
        #기록 엑셀에 저장
        
        
        for i in range(0, len(name)):
                sheet.cell(row = 2+i+accumulate_sum, column = 1).value = name[i] # 선수명 기입
                sheet.cell(row = 2+i+accumulate_sum, column = 2).value = backnum[i] # 등번호 기입
                sheet.cell(row = 2+i+accumulate_sum, column = 3).value = Player_Batter_Stats[i]['playerId']
                sheet.cell(row = 2+i+accumulate_sum, column = 4).value = year
                sheet.cell(row = 2+i+accumulate_sum, column = 5).value = Player_Batter_Stats[i]['G']
                sheet.cell(row = 2+i+accumulate_sum, column = 6).value = Player_Batter_Stats[i]['PA']
                sheet.cell(row = 2+i+accumulate_sum, column = 7).value = Player_Batter_Stats[i]['AB']
                sheet.cell(row = 2+i+accumulate_sum, column = 8).value = Player_Batter_Stats[i]['Runs']
                sheet.cell(row = 2+i+accumulate_sum, column = 9).value = Player_Batter_Stats[i]['Hits']
                sheet.cell(row = 2+i+accumulate_sum, column = 10).value = Player_Batter_Stats[i]['_2B']
                sheet.cell(row = 2+i+accumulate_sum, column = 11).value = Player_Batter_Stats[i]['_3B']
                sheet.cell(row = 2+i+accumulate_sum, column = 12).value = Player_Batter_Stats[i]['HR']
                sheet.cell(row = 2+i+accumulate_sum, column = 13).value = Player_Batter_Stats[i]['TB']
                sheet.cell(row = 2+i+accumulate_sum, column = 14).value = Player_Batter_Stats[i]['RBI']
                sheet.cell(row = 2+i+accumulate_sum, column = 15).value = Player_Batter_Stats[i]['SB']
                sheet.cell(row = 2+i+accumulate_sum, column = 16).value = Player_Batter_Stats[i]['CS']
                sheet.cell(row = 2+i+accumulate_sum, column = 17).value = Player_Batter_Stats[i]['BB']
                sheet.cell(row = 2+i+accumulate_sum, column = 18).value = Player_Batter_Stats[i]['IBB']
                sheet.cell(row = 2+i+accumulate_sum, column = 19).value = Player_Batter_Stats[i]['HBP']
                sheet.cell(row = 2+i+accumulate_sum, column = 20).value = Player_Batter_Stats[i]['SO']
                sheet.cell(row = 2+i+accumulate_sum, column = 21).value = Player_Batter_Stats[i]['GDP']
                sheet.cell(row = 2+i+accumulate_sum, column = 22).value = Player_Batter_Stats[i]['SAC']
                sheet.cell(row = 2+i+accumulate_sum, column = 23).value = Player_Batter_Stats[i]['SF']
                sheet.cell(row = 2+i+accumulate_sum, column = 24).value = Player_Batter_Stats[i]['AVG']
                sheet.cell(row = 2+i+accumulate_sum, column = 25).value = Player_Batter_Stats[i]['OBP']
                sheet.cell(row = 2+i+accumulate_sum, column = 26).value = Player_Batter_Stats[i]['SLG']
                sheet.cell(row = 2+i+accumulate_sum, column = 27).value = Player_Batter_Stats[i]['OPS']
                sheet.cell(row = 2+i+accumulate_sum, column = 28).value = Player_Batter_Stats[i]['WAR']
                sheet.cell(row = 2+i+accumulate_sum, column = 29).value = Player_Batter_Stats[i]['GO']
                sheet.cell(row = 2+i+accumulate_sum, column = 30).value = Player_Batter_Stats[i]['FO']
                sheet.cell(row = 2+i+accumulate_sum, column = 31).value = Player_Batter_Stats[i]['LO']
                sheet.cell(row = 2+i+accumulate_sum, column = 32).value = Player_Batter_Stats[i]['P']
                sheet.cell(row = 2+i+accumulate_sum, column = 33).value = Player_Batter_Stats[i]['C']
                sheet.cell(row = 2+i+accumulate_sum, column = 34).value = Player_Batter_Stats[i]['1B']
                sheet.cell(row = 2+i+accumulate_sum, column = 35).value = Player_Batter_Stats[i]['2B']
                sheet.cell(row = 2+i+accumulate_sum, column = 36).value = Player_Batter_Stats[i]['3B']
                sheet.cell(row = 2+i+accumulate_sum, column = 37).value = Player_Batter_Stats[i]['SS']
                sheet.cell(row = 2+i+accumulate_sum, column = 38).value = Player_Batter_Stats[i]['LF']
                sheet.cell(row = 2+i+accumulate_sum, column = 39).value = Player_Batter_Stats[i]['LC']
                sheet.cell(row = 2+i+accumulate_sum, column = 40).value = Player_Batter_Stats[i]['CF']
                sheet.cell(row = 2+i+accumulate_sum, column = 41).value = Player_Batter_Stats[i]['RC']
                sheet.cell(row = 2+i+accumulate_sum, column = 42).value = Player_Batter_Stats[i]['RF']

        accumulate_sum += len(name)        
        print(team_name[0].text + " (" + club_idt[teams] + ") 기록 완료!!!!!")
        #엑셀 파일 만들기
        excel_file.save(path2 + '[분석] '+ year + ' AUBL.xlsx')
        #excel_file.save(path + '[일괄분석] '+ year + ' ' + team_name[0].text +'.xlsx')
        #excel_file.close()
        
#excel_file.save(path2 + '[분석] '+ year + ' AUBL.xlsx')
excel_file.close()
        

        
