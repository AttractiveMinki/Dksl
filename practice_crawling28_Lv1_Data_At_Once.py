import bs4
import urllib.request
import os
import openpyxl
import time

#연도 입력
year = '2019'
club_id = '25226' #

club_idt = [0 for i in range(40)]
club_idt[0] = '32122'#한체대
club_idt[1] = '33835'#가천대
club_idt[2] = '23604'#가톨릭대
club_idt[3] = '11841'#건대글로컬
club_idt[4] = '11866'#건대서울
club_idt[5] = '24037'#경기대
club_idt[6] = '23749'#경희국제
club_idt[7] = '23553'#경희서울
club_idt[8] = '11853'#고려대
club_idt[9] = '23999'#광운대
club_idt[10] = '11714'#국민대
club_idt[11] = '21584'#단국죽전
club_idt[12] = '11869'#단국천안
club_idt[13] = '23987'#동국대
club_idt[14] = '11849'#명지서울
club_idt[15] = '23995'#명지용인
club_idt[16] = '25226'#백석대
club_idt[17] = '23515'#상명대
club_idt[18] = '11721'#서강대
club_idt[19] = '23686'#서경대
club_idt[20] = '970'#서울과학기술대
club_idt[21] = '11709'#서울시립대
club_idt[22] = '24158'#성균관대
club_idt[23] = '11810'#세종대
club_idt[24] = '11877'#숭실대
club_idt[25] = '343'#아주대
club_idt[26] = '11919'#연세대
club_idt[27] = '11847'#외대글로벌
club_idt[28] = '23676'#인천대
club_idt[29] = '471'#인하대
club_idt[30] = '24016'#중앙대
club_idt[31] = '28867'#한국교통대
club_idt[32] = '6205'#산기대
club_idt[33] = '23728'#한국외대서울
club_idt[34] = '23593'#항공대
club_idt[35] = '24010'#한성대
club_idt[36] = '23691'#한신대
club_idt[37] = '24012'#한양대에리카
club_idt[38] = '23961'#한양대서울
club_idt[39] = '23511'#홍익대
for teams in range(0, len(club_idt)):
        url = "http://www.gameone.kr/club/info/schedule/table?club_idx=" + club_idt[teams] + "&kind=&season=" + year # 한체대 주소
        #향후 32122 이 부분을 한체대에 배정할 예정 ex) 한체대 = 32122
        html = urllib.request.urlopen(url)
        bsObj = bs4.BeautifulSoup(html, "html.parser")

        #위 공백 갯수 입력
        space = 5

        #팀 이름
        team_name = bsObj.select('title') # team_name[0].text 로 사용

        #경로 지정
       # path = "/Users/YEO/AppData/Local/Programs/Python/Python38/"+ year + "/" + team_name[0].text + "/"
        path = "/Users/JuYeop/AppData/Local/Programs/Python/Python38-32/"+ year + "/" + team_name[0].text + "/"

        #폴더 조회 후 없으면 만들기
        #import os

        if(os.path.exists(path) != True): # 폴더가 존재하지 않는다면
                os.makedirs(path) # 폴더를 만든다.

        #import openpyxl
        excel_file = openpyxl.Workbook()
        excel_file.remove(excel_file.active)
        excel_sheet = excel_file.create_sheet('선수 이름')

        page=1

        #팀 경기 기록 주소
        team_record = bsObj.select('td > a') 

        #팀 경기 기록 주소 할당
        team_address = [0 for i in range(len(team_record))]

        for i in range(0, len(team_record)):
                team_address[i] = team_record[i].get('href')
                i+=1

        url_initial = 'https://www.gameone.kr'

        #https 포함 주소 할당해주기
        game_url = [0 for i in range(len(team_address))]

        i=0
        for i in range(0, len(team_address)):
                game_url[i] = url_initial + team_address[i]
                i+=1


        ## 경기 기록 추가 탐색
        flag = True
        while(flag):
            if(len(team_record) < 15):
                flag = False
                break
            page+=1
            page_str = str(page)
            newurl = url + '&game_type=0&lig_idx=0&group=0&month=0&page=' + page_str
            html = urllib.request.urlopen(newurl)
            bsObj = bs4.BeautifulSoup(html, "html.parser")
                    
            #import openpyxl
            excel_file = openpyxl.Workbook()
            excel_file.remove(excel_file.active)
            excel_sheet = excel_file.create_sheet('팀 기록1')
                    
            # 팀 경기 기록 주소
            team_record = bsObj.select('td > a')
            count = (len(team_record) * (page-1))

            #팀 경기 기록 주소 할당
            team_address_add = [0 for i in range(len(team_record))]

            for i in range(0, len(team_record)):
                team_address_add[i] = team_record[i].get('href')
                i+=1

            team_address.extend(team_address_add)

            url_initial = 'https://www.gameone.kr'

            #https포함 주소 할당해주기
            game_url_add = [0 for i in range(len(team_record))]

            i=0
            for i in range(0, len(team_record)):
                game_url_add[i] = url_initial + team_address_add[i]
                i+=1

            game_url.extend(game_url_add)

        ## url 바탕으로 기록 저장
        for number_of_times in range(0, len(game_url)):
            url = game_url[number_of_times]
            html = urllib.request.urlopen(url)
            bsObj = bs4.BeautifulSoup(html, "html.parser")
            see_resultAll = bsObj.find("div", {"class": "record"})

            #팀 이름
            NameAll = see_resultAll.find_all("span", {"class" : "name"}) # 선수 이름
            Batter_record_All = see_resultAll.find_all("td", {"class" : "round"}) # 타자 기록
            Batting_order = see_resultAll.find_all("span", {"class" : "num"}) # 선수 타순, 원정 타자 수 + 홈 타자 수로 변경 가능
            Pitcher_record_form = see_resultAll.select('div.record > table.record_table > thead > tr > th') #투수 기록 양식
            see_Pitcher = bsObj.select('tr > th > strong') # 투수 이름 가져오기
            see_result_Pitch = bsObj.select('td') #투수 기록 가져오기

            excel_file = openpyxl.Workbook()
            excel_file.remove(excel_file.active)
            excel_sheet = excel_file.create_sheet('팀 기록1')

            sheet = excel_file.active

            #날짜 저장
            Game_info = bsObj.select('div.section_score > table > caption')
            Game_date1 = Game_info[0].text.strip().replace("/","")
            Game_date2 = Game_date1.strip().replace(":","_")
            Game_date3 = Game_date2.strip().replace("<","")
            Game_date = Game_date3.strip().replace(">","")
            ########################

            #타자 기록 양식
            Record_top = see_resultAll.select('div.record > table:nth-child(2) > thead > tr > th')

            #원정 타자 이름
            Batter_name_away = see_resultAll.select('div.record > table:nth-child(2) > tbody > tr > th > span')
            #홈 타자 이름
            Batter_name_home = see_resultAll.select('div.record > table:nth-child(4) > tbody > tr > th > span')

            #원정 타자 수
            Batter_count_away = see_resultAll.select('div.record > table:nth-child(2) > tbody > tr > th > span > strong')
            #홈 타자 수
            Batter_count_home = see_resultAll.select('div.record > table:nth-child(4) > tbody > tr > th > span > strong')

            #원정 타자 기록
            Batter_record_away = see_resultAll.select('div.record > table:nth-child(2) > tbody > tr > td')
            #홈 타자 기록
            Batter_record_home = see_resultAll.select('div.record > table:nth-child(4) > tbody > tr > td')

            #타자 기록 양식
            Record_num = [0 for i in range(len(Record_top))]

            num=0
            for record in range(0, len(Record_top)):
                    Record_num[num] = Record_top[num].text.strip().replace("\'","")
                    num+=1

            #타자 기록 양식 엑셀에 입력
            count = 0
            num=0
            inning = 0
            for count in range(0, len(Record_top)):
                sheet.cell(row = 1 + space, column = count + 3).value = Record_num[num]
                num+=1
                   
            #원정 타자 이름
            Batting_array_away = [0 for i in range(len(Batter_name_away))]

            num=0
            for batter in range(0, len(Batter_name_away)):
                    Batting_array_away[num] = Batter_name_away[num].text.strip().replace("\'","")
                    num+=1
                    
            #원정 타자 이름 엑셀에 입력
            count = 0
            num=0
            inning = 0
            for count in range(0, len(Batter_count_away)):
                for inning in range(1, 4):
                   sheet.cell(row = 2 + count + space, column = inning).value = Batting_array_away[num]
                   num+=1


            #홈 타자 이름
            Batting_array_home = [0 for i in range(len(Batter_name_home))]

            num=0
            for batter in range(0, len(Batter_name_home)):
                    Batting_array_home[num] = Batter_name_home[num].text.strip().replace("\'","")
                    num+=1
                    
            #홈 타자 이름 엑셀에 입력
            count = 0
            num=0
            inning = 0
            for count in range(0, len(Batter_count_home)):
                for inning in range(1, 4):
                   sheet.cell(row = 3 + count + len(Batter_count_away) + space, column = inning).value = Batting_array_home[num]
                   num+=1


            #원정 타자 기록
            Batting_order_away = [0 for i in range(len(Batter_count_away)*19)]

            num=0
            for batter in range(0, len(Batter_record_away)):
                    Batting_order_away[num] = Batter_record_away[num].text.strip().replace("\'","")
                    num+=1

            #원정 타자 기록 엑셀에 입력
            count = 0
            num=0
            inning = 0
            for count in range(0, len(Batter_count_away)):
                for inning in range(0, 12+7):
                   sheet.cell(row = 2 + count + space, column = 4 + inning).value = Batting_order_away[num]
                   num+=1


            #홈 타자 기록
            Batting_order_home = [0 for i in range(len(Batter_count_home)*19)]

            num=0
            for batter in range(0, len(Batter_record_home)):
                    Batting_order_home[num] = Batter_record_home[num].text.strip().replace("\'","")
                    num+=1

            #홈 타자 기록 엑셀에 입력
            count = 0
            num=0
            inning = 0
            for count in range(0, len(Batter_count_home)):
                for inning in range(0, 12+7):
                   sheet.cell(row = count + 3 + len(Batter_count_away) + space, column = 4 + inning).value = Batting_order_home[num]
                   num+=1


            #선발투수 구별하기
            see_result_p = bsObj.select('tbody > tr > th > strong') # 양 팀 선발투수 가져오기
            sequence = 0
            for i in range(0, len(see_Pitcher)):
                if(len(see_result_p) < 2):#한 쪽 선발투수만 입력된 경우 오류 방지
                        sequence = 0 # 원정팀 등판 투수 숫자 = sequence
                        break
                elif(see_result_p[1].text == see_Pitcher[i].text): # 홈 팀 선발투수 이름 가져오기
                        sequence = i # 원정팀 등판 투수 숫자 = sequence
                        break

            #원정팀 투수 이름 저장
            see_Pitcher_away = [0 for i in range(sequence)]

            num = 0
            for record in range(0, sequence):
                see_Pitcher_away[num] = see_Pitcher[num].text.strip().replace("\'","")
                num+=1

            #홈팀 투수 이름 저장
            see_Pitcher_home = [0 for i in range(len(see_Pitcher)-sequence)]

            num = 0
            for record in range(0, len(see_Pitcher)-sequence):
                see_Pitcher_home[num] = see_Pitcher[sequence + num].text.strip().replace("\'","")
                num+=1

            #원정팀 투수 이름 입력
            count = 0
            num=0
            for count in range(0, sequence): # Pitcher 수로 수정예정, 이 경기는 6명. 투수 이름 가져온 뒤 len으로 수정!!
                sheet.cell(row = 5 + len(NameAll) + count + space, column = 1).value = see_Pitcher_away[num]
                num+=1

            #홈팀 투수 이름 입력
            count = 0
            num=0
            for count in range(0, len(see_Pitcher)-sequence): # Pitcher 수로 수정예정, 이 경기는 6명. 투수 이름 가져온 뒤 len으로 수정!!
                sheet.cell(row = 6 + len(NameAll) + sequence + count + space, column = 1).value = see_Pitcher_home[num]
                num+=1

                
            #투구 기록 양식
            Pitching_record_help = [0 for i in range(19)]
            #range의 경우 len(Pitcher_record_form)-40 하면 좋은데, 한 팀의 투수가 없을 경우 error

            num=0
            for record in range(0, len(Pitching_record_help)):
                Pitching_record_help[num] = Pitcher_record_form[num+40].text.strip().replace("\'","")
                num+=1
                
            count = 0
            num = 0
            for count in range(0, len(Pitching_record_help)):
                    sheet.cell(row = 4 + len(NameAll) + space, column = 1+count).value = Pitching_record_help[num]
                    num+=1


            minus = (len(Batting_order)+1)*19+27; # 12회까지 12개, 기록 관련 7개, 21*19, 위쪽 경기 스코어 26
            Pitching_record_support = [0 for i in range(len(see_result_Pitch)-minus)]


            num=0
             #타자기록 도움 항목 제외
            for record in range(0, len(Pitching_record_support)):
                Pitching_record_support[num] = see_result_Pitch[num+minus].text.strip().replace("\'","")
                num+=1

            #원정팀 투수 기록 저장
            Pitching_record_away = [0 for i in range(sequence*(len(Pitching_record_help)-1))]

            num=0
            for num in range(0, sequence*(len(Pitching_record_help)-1)):
                    Pitching_record_away[num] = Pitching_record_support[num]
                    num+=1

            #홈팀 투수 기록 저장
            Pitching_record_home = [0 for i in range((len(see_Pitcher)-sequence)*(len(Pitching_record_help)-1))]

            num=0
            for num in range(0, (len(see_Pitcher)-sequence)*(len(Pitching_record_help)-1)):
                    Pitching_record_home[num] = Pitching_record_support[num + ((len(Pitching_record_help)-1) * sequence)]
                    num+=1


            #원정팀 투수 기록 입력
            count = 0
            num=0
            for count in range(0, sequence): # Pitcher 수로 수정예정, 이 경기는 6명. 투수 이름 가져온 뒤 len으로 수정!!
                for record in range(0,18):
                    sheet.cell(row = 5 + len(NameAll) + count + space, column = 2 + record).value = Pitching_record_away[num]
                    num+=1

            #홈팀 투수 기록 입력
            count = 0
            num=0
            for count in range(0, len(see_Pitcher)-sequence): # Pitcher 수로 수정예정, 이 경기는 6명. 투수 이름 가져온 뒤 len으로 수정!!
                for record in range(0,18):
                    sheet.cell(row = 6 + len(NameAll) + sequence + count + space, column = 2 + record).value = Pitching_record_home[num]
                    num+=1

                    
            #원정팀, 홈팀 저장
            Team_play = bsObj.select('div.record > h3')
            Team_away = Team_play[0].text
            Team_home = Team_play[1].text
        
            if(Team_away.find("*") != -1):
                    Team_away = Team_away.replace("*","_")

            if(Team_home.find("*") != -1):
                    Team_home = Team_home.replace("*","_")        
                
            #원정팀 입력
            sheet.cell(row = 1 + space, column = 2).value = Team_away # 공격 기록 위
            sheet.cell(row = 3+len(NameAll) + space, column = 2).value = Team_away # 투수 기록 위

            #홈팀 입력
            sheet.cell(row = 2 + len(Batter_count_away) + space, column = 2).value = Team_home # 공격 기록 위
            sheet.cell(row = 5+len(NameAll)+sequence + space, column = 2).value = Team_home # 투수 기록 위

            cell_A1 = excel_sheet['A1']
            cell_A1.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell_B1 = excel_sheet['B1']
            cell_B1.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell_C1 = excel_sheet['C1']
            cell_C1.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell_D1 = excel_sheet['D1']
            cell_D1.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell_E1 = excel_sheet['E1']
            cell_E1.alignment = openpyxl.styles.Alignment(horizontal='center')
            print(team_name[0].text + " " + str(number_of_times) + "번째 기록 크롤링중")
            excel_file.save(path + Team_away + ' vs ' + Team_home + ' ' + Game_date +'.xlsx')
            excel_file.close()
            #import time
            time.sleep(0.2)
            if(number_of_times > 5):
                    time.sleep(0.2) # 시간 딜레이 1초
                    if(number_of_times > 10):
                            time.sleep(0.2)
                            if(number_of_times > 15):
                                    time.sleep(0.3)
                                    if(number_of_times > 20):
                                            time.sleep(0.1)
        print(team_name[0].text + "기록 크롤링 완료!!!!!")
#excel_file.save(team_name[0].text + '_' + year + '_' + i + '.xlsx')
#excel_file.close()

